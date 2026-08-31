"""组件测试：python tests/test_components.py 直接运行（T1）。"""

import io
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import pandas as pd
from openpyxl import load_workbook


def test_excel_bytes():
    from components.export import excel_bytes
    df = pd.DataFrame({"a": [1, 2], "b": ["x", "y"]})
    data = excel_bytes({"表一": df, "表二": df})
    wb = load_workbook(io.BytesIO(data))
    assert wb.sheetnames == ["表一", "表二"]
    assert wb["表一"].max_row == 3  # 表头 + 2 行


def test_pdf_bytes():
    from components.export import pdf_bytes
    data = pdf_bytes("测试报告", [("指标甲", "100"), ("指标乙", "20%")], ["中文要点一", "要点二"])
    assert data[:5] == b"%PDF-"
    assert len(data) > 1000


def test_apptest_home():
    from streamlit.testing.v1 import AppTest
    at = AppTest.from_file(str(Path(__file__).resolve().parent.parent / "app.py"),
                           default_timeout=30)
    at.run()
    assert not at.exception, f"首页异常: {at.exception}"


if __name__ == "__main__":
    test_excel_bytes()
    test_pdf_bytes()
    test_apptest_home()
    print("ALL COMPONENT TESTS PASS")
